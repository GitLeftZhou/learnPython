# -*- coding:utf-8 -*-
import configparser
import os
import sys

import openpyxl
from openpyxl.cell.cell import Cell
from openpyxl.styles import Font
from openpyxl.styles import colors, PatternFill
from openpyxl.worksheet import Worksheet


class Compare:

    def __init__(self, work_dir=None):
        """
        初始化，加载配置文件信息到内存中
        """
        # 获取当前目录
        if work_dir is None:
            self.base_dir = os.path.dirname(os.path.realpath(sys.argv[0]))
        else:
            self.base_dir = work_dir
        print("\n=======================================================")
        print("script running at " + self.base_dir)
        print("=======================================================")
        # 读取配置文件
        self.cf = configparser.ConfigParser()
        if os.path.exists("configuration.config"):
            self.cf.read("configuration.config")[0]
            # 加载EXCEL
            if self.cf.get("excel", "templet_file") is not None:
                self.templet_file = os.path.join(self.base_dir, self.cf.get("excel", "templet_file"))
            else:
                self.templet_file = None

            if self.cf.get("excel", "target_file") is not None:
                self.tgt_file = os.path.join(self.base_dir, self.cf.get("excel", "target_file"))
            else:
                self.tgt_file = None
            # 获取对比列
            self.templet_match_col = self.cf.get("compare", "templet_match_col")  # 根据S列匹配
            self.templet_compare_col = self.cf.get("compare", "templet_compare_col")  # 比较U列值
            self.target_match_col = self.cf.get("compare", "target_match_col")  # 根据S列匹配
            self.target_compare_col = self.cf.get("compare", "target_compare_col")  # 比较U列值
        else:
            self.cf = None
            print("没有找到configuration.config配置文件")

    # 进度条
    def report(self, count, current, job):
        percent = int(current * 100 / count)
        sys.stdout.write("\r%s has been completed %d%% " % (job, percent))
        sys.stdout.flush()

    def cache_template(self):
        """
        缓存模版表数据到字典中
        :return:
        """
        templet_dict = {}
        if os.path.exists(self.templet_file):
            base_xlsx = openpyxl.load_workbook(self.templet_file, read_only=True)

            # 读取模版表到字典中
            templet_sheet: Worksheet = base_xlsx[base_xlsx.sheetnames[0]]
            templet_max_row = templet_sheet.max_row
            print("模版文件中检索到%d条记录\n" % templet_max_row)
            for i in range(1, templet_max_row + 1):
                self.report(templet_max_row, i, "cache templet ")
                identity = str(templet_sheet[self.templet_match_col + str(i)].value)
                tgt_value = str(templet_sheet[self.templet_compare_col + str(i)].value)
                if identity is not None:
                    # 忽略大小写
                    identity: str = identity.upper()
                    tgt_value: str = tgt_value.upper()
                    # 替换掉特殊字符 只比较字母
                    identity = identity.replace(".", "")
                    identity = identity.replace(",", "")
                    identity = identity.replace(" ", "")
                    identity = identity.replace("/", "")
                    identity = identity.replace("\\", "")
                    identity = identity.replace("(", "")
                    identity = identity.replace(")", "")
                    identity = identity.replace("\'", "")
                    if identity != "NONE":
                        templet_dict.update({identity: tgt_value})
            base_xlsx.close()  # 关闭模版文件
            print("缓存模版文件完成.共缓存%d条记录\n" % len(templet_dict))

        else:
            print("没有找到模版文件")

        return templet_dict

    def compare(self, tgt_file, templet_dict):
        """

        :param tgt_file: 目标文件
        :param templet_dict: 数据字典
        :return:
        """
        tgt_xlsx = openpyxl.load_workbook(tgt_file)
        # 读取目标文件
        target_sheet: Worksheet = tgt_xlsx[tgt_xlsx.sheetnames[0]]
        target_max_row = target_sheet.max_row
        print("目标文件中检索到%d条记录\n" % target_max_row)
        print("=====================================================================")
        print("\n处理目标：匹配列不存在，设置背景色为黄色。对比列不同，设置文字为红色\n")
        print("=====================================================================")
        # 处理目标文件
        for i in range(1, target_max_row + 1):
            self.report(target_max_row, i, "compare file ")
            match_cell: Cell = target_sheet[self.target_match_col + str(i)]
            compare_cell: Cell = target_sheet[self.target_compare_col + str(i)]
            identity = str(match_cell.value)
            tgt_value = str(compare_cell.value)

            if identity is not None:
                # 忽略大小写
                identity = identity.upper()
                tgt_value = tgt_value.upper()
                # 替换掉特殊字符 只比较字母
                identity = identity.replace(".", "")
                identity = identity.replace(",", "")
                identity = identity.replace(" ", "")
                identity = identity.replace("/", "")
                identity = identity.replace("\\", "")
                identity = identity.replace("(", "")
                identity = identity.replace(")", "")
                identity = identity.replace("\'", "")
                if identity != "NONE":
                    if identity in templet_dict:
                        if tgt_value != templet_dict.get(identity):
                            print("处理第%d条记录，匹配列: %s" % (i, identity), end="  =  ")
                            print(tgt_value, end="")
                            print("  处理方式：模版文件匹配列的对应值不相等，设置文字为红色")
                            ft = Font(color=colors.RED)
                            compare_cell.font = ft
                    else:
                        print("处理第%d条记录，匹配列: %s" % (i, identity), end="  =  ")
                        print(tgt_value, end="")
                        print("  处理方式：模版文件中没有这个匹配值，设置背景为黄色")
                        match_cell.fill = PatternFill(fill_type='solid', fgColor=colors.YELLOW)
        new_file = os.path.splitext(tgt_file)
        tgt_xlsx.save(str(new_file[0])+str("(1)")+str(new_file[1]))
        tgt_xlsx.close()


# 运行入口
if __name__ == "__main__":

    try:
        base_dir = os.path.dirname(os.path.realpath(sys.argv[0]))
        compare = Compare(base_dir)

        if compare.cf is not None:
            templet_dict = compare.cache_template()
            if compare.tgt_file is None or not os.path.exists(compare.tgt_file):
                for tgt_file in os.listdir(base_dir):
                    full_file_path = os.path.join(base_dir, tgt_file)
                    if not os.path.isdir(full_file_path):
                        if (tgt_file.endswith("xlsx") or tgt_file.endswith("xls")) and not compare.templet_file == full_file_path:
                            compare.compare(full_file_path, templet_dict)
            else:
                compare.compare(compare.tgt_file,templet_dict)

        print("\n=======================================================")
        print("=======================================================")
        input("按回车键结束  ")

    except BaseException as e:
        print(e)
        input("按回车键退出程序  ")


