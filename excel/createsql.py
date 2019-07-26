# -*- coding:utf-8 -*-
import sys

import openpyxl
from openpyxl.worksheet import Worksheet

# from openpyxl.cell.cell import Cell

# 加载EXCEL
wb = openpyxl.load_workbook("data.xlsx")


# sheet_names = wb.get_sheet_names()
# print(sheetnames)
# sheet_menu: Worksheet = wb["目录"]
# cell_b1 = sheet_menu.cell(row=1,column=2) 根据行号，列号获取单元格
# cell_b1 = sheet_menu["B1"]
# cell_b1 = sheet_menu.cell("B1")
# print(cell_b1.value)
# sheet_columns = sheet_menu.columns
# max_row = sheet_menu.max_row
# min_row = sheet_menu.min_row
# print("max_row:%d,min_row:%d" % (max_row, min_row))
# 获取多个单元格
# cells = sheet_menu['C3:C' + str(max_row)]
# print("cells is %s" % (str(type(cells))))
# 遍历多个单元格 需循环两次

# 进度条
def report(count, current, job):
    percent = int(current * 100 / count)
    sys.stdout.write("\r%s has been completed %d%% " % (job, percent))
    sys.stdout.flush()


# for sheetName in wb.sheetnames:
sheetName = wb.sheetnames[1]
current_sheet: Worksheet = wb[sheetName]
max_idx = current_sheet.max_row + 1
print(sheetName + ":" + str(max_idx))
start_str = r"INSERT INTO BDDJ3_CHECK_RESULT SELECT '"
split_str = r"','"
with open("data.txt", mode='a', encoding="utf-8") as f:
    for i in range(1, max_idx):
        rule_no = str(current_sheet["A" + str(i)].value)
        table_name = str(current_sheet["B" + str(i)].value)
        col_name = str(current_sheet["C" + str(i)].value)
        rule_name = str(current_sheet["D" + str(i)].value)
        sql_text = (start_str + rule_no + split_str + table_name + split_str + col_name
                    + split_str + rule_name + r"', A.Transactionno"
                    + r",'H' FROM " + table_name
                    + " A WHERE A.Batchno IN (SELECT BATCHNO FROM BDDJ3_CHECK_BATCHNO) \r\n; ")
        # print(sql_text)
        f.write(sql_text + "\r\n\r\n")
        # f.write("\r\n")
        report(max_idx, i, sheetName+"....")
wb.close()
