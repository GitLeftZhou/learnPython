# -*- coding:utf-8 -*-
import openpyxl, logging, sys
from openpyxl.worksheet import Worksheet
from openpyxl.cell.cell import Cell

# 注册日志 filename='log.log'
logging.basicConfig(filename='log.log', level=logging.DEBUG, format=" %(asctime)s - %(levelname)s %(message)s")

# 加载EXCEL
wb = openpyxl.load_workbook("codedata.xlsx")
# 打开文件
txt_file = open("data.txt", "a")
# sheet_names = wb.get_sheet_names()
# print(sheetnames)
# sheet_menu: Worksheet = wb["目录"]
# cell_b1 = sheet_menu.cell(row=1,column=2) 根据行号，列号获取单元格
# cell_b1 = sheet_menu["B1"]
# cell_b1 = sheet_menu.cell("B1")
# print(cell_b1.value)
# sheet_columns = sheet_menu.columns
# max_row = sheet_menu.max_row
# min_row = sheet_menu.min_row
# print("max_row:%d,min_row:%d" % (max_row, min_row))
# 获取多个单元格
# cells = sheet_menu['C3:C' + str(max_row)]
# print("cells is %s" % (str(type(cells))))
# 遍历多个单元格 需循环两次


for sheet_name in wb.sheetnames:

    try:
        logging.info(type(sheet_name))
        logging.info(sheet_name)
        logging.info(type(sheet_name))
        logging.info("sheet_name = %s"%sheet_name)
        if sheet_name not in ['封面', '目录']:
            curr_sheet: Worksheet = wb[sheet_name]
            result_dict = {}
            for sheet_columns in curr_sheet.columns:
                columns_lst = list(tuple(sheet_columns))
                tmp_list = []
                for curr_cell in columns_lst:
                    bottom_border = curr_cell.border.bottom.style
                    top_border = curr_cell.border.top.style
                    # 获取有效表格
                    if bottom_border is not None and top_border is not None:
                        # 排除头
                        sex_cell_value = str(curr_cell.value)
                        if sheet_name == sex_cell_value.replace('/', ''):
                            continue
                        tmp_list.append(curr_cell.value)
                    else:
                        continue
                if tmp_list:
                    result_dict[tmp_list.pop(0)] = tmp_list
            logging.info(result_dict)
            # 此处可以组装脚本写入文件
            # txt_file.write()
    except BaseException as e:
        logging.info(e)

txt_file.close()
wb.close()
